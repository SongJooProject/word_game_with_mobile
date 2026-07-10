"""
엑셀 문제 데이터를 JSON으로 변환하는 스크립트
규칙:
  - type1 (선택형): D열에 선택지 있음
  - type2 (빈칸형): 문제에 < > 있음
  - 섹션 헤더: B열/선택지/정답 모두 없음 → 제외
"""

import openpyxl
import json
import re
from pathlib import Path


def detect_type(row):
    """문제 유형 판별"""
    col_b = row[1]   # type 열
    col_c = row[2]   # 문제 텍스트
    col_d = row[3]   # 선택지1

    # B열에 type이 명시되어 있으면 사용
    if col_b in ["type1", "type2"]:
        return col_b

    # D열에 선택지 있으면 type1
    if col_d is not None:
        return "type1"

    # 문제에 < > 있으면 type2
    if col_c and "<" in str(col_c) and ">" in str(col_c):
        return "type2"

    # 그 외: 섹션 헤더이거나 문제 아님
    return None


def parse_excel_to_questions(excel_path):
    """엑셀 파일을 읽어 문제 데이터로 변환"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    all_chapters = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        questions = []
        current_section = ""

        rows = ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
        for row_idx, row in enumerate(rows, 1):
            cells = list(row)

            # 빈 행 처리
            if all(v is None for v in cells):
                continue

            col_b = cells[1] if len(cells) > 1 else None
            col_c = cells[2] if len(cells) > 2 else None

            # 문제 유형 판별
            q_type = detect_type(cells)

            # 섹션 헤더 감지 (B열 없고, D열 없고, < >도 없음)
            if q_type is None and col_b is None and col_c is not None:
                current_section = str(col_c).strip()
                continue

            if q_type is None:
                continue

            content = str(col_c).strip() if col_c else ""
            if not content:
                continue

            if q_type == "type1":
                # 선택형 문제
                options = []
                for i in range(3, 9):  # D~I열
                    if i < len(cells) and cells[i] is not None:
                        options.append(str(cells[i]).strip())

                answer_raw = cells[10] if len(cells) > 10 else None
                if answer_raw and str(answer_raw).isdigit():
                    answer = int(answer_raw)
                else:
                    answer = 1

                if options:
                    questions.append({
                        "type": "type1",
                        "section": current_section,
                        "question": content,
                        "options": options,
                        "answer": answer
                    })

            elif q_type == "type2":
                # 빈칸형 문제
                blanks = re.findall(r"<([^>]+)>", content)
                answer = "|".join(blanks) if blanks else ""

                questions.append({
                    "type": "type2",
                    "section": current_section,
                    "question": content,
                    "answer": answer
                })

        if questions:
            all_chapters.append({
                "name": sheet_name,
                "questions": questions
            })

    return all_chapters


def main():
    excel_path = Path("db/DB_수사종결까지.xlsx")
    output_path = Path("data/questions.json")

    print("=" * 50)
    print("엑셀 → JSON 변환")
    print("=" * 50)

    if not excel_path.exists():
        print(f"에러: {excel_path} 파일을 찾을 수 없습니다.")
        return

    chapters = parse_excel_to_questions(excel_path)

    # JSON 구조 생성
    data = {
        "subjects": [
            {
                "name": "형사소송법",
                "chapters": chapters
            }
        ]
    }

    # 저장
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    # 통계
    total_q = sum(len(ch["questions"]) for ch in chapters)
    print("\n변환 완료!")
    print(f"  - 시트(챕터): {len(chapters)}개")
    print(f"  - 전체 문제: {total_q}개")

    for ch in chapters:
        type1_count = sum(1 for q in ch["questions"] if q["type"] == "type1")
        type2_count = sum(1 for q in ch["questions"] if q["type"] == "type2")
        print(f"\n  [{ch['name']}]")
        print(f"    - 선택형: {type1_count}개")
        print(f"    - 빈칸형: {type2_count}개")

    print(f"\n저장: {output_path}")


if __name__ == "__main__":
    main()
