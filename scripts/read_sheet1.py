import openpyxl

wb = openpyxl.load_workbook("db/DB....xlsx", data_only=True)
ws = wb["Sheet1"]

print(f"Sheet1 전체 행 수: {ws.max_row}")
print(f"Sheet1 전체 열 수: {ws.max_column}")
print()

for row_idx, row in enumerate(
    ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False),
    1,
):
    print(f"=== {row_idx}행 ===")
    for cell in row:
        if cell.value is not None:
            print(f"  {cell.coordinate}: {cell.value}")
    print()
