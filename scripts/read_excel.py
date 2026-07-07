import openpyxl
import json

wb = openpyxl.load_workbook("db/DB....xlsx", data_only=True)

result = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    result.append(f"\n=== {sheet_name} ===")
    result.append(f"행 수: {ws.max_row}, 열 수: {ws.max_column}\n")
    
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=False):
        row_data = []
        for cell in row:
            if cell.value is not None:
                row_data.append(f"{cell.coordinate}: {cell.value}")
        if row_data:
            result.append(" | ".join(row_data))

with open("db/excel_read.txt", "w", encoding="utf-8") as f:
    f.write("\n".join(result))
