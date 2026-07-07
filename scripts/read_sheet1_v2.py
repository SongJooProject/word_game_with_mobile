import openpyxl

wb = openpyxl.load_workbook("db/DB....xlsx", data_only=True)
ws = wb["Sheet1"]

with open("db/sheet1_detail.txt", "w", encoding="utf-8") as f:
    f.write(f"Sheet1: {ws.max_row}행 x {ws.max_column}열\n\n")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
        f.write(f"=== {row_idx}행 ===\n")
        for cell in row:
            if cell.value is not None:
                f.write(f"  {cell.coordinate}: {cell.value}\n")
        f.write("\n")
