import openpyxl
import json

wb = openpyxl.load_workbook("db/DB....xlsx", data_only=True)

result = {}
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        rows.append([str(cell) if cell is not None else "" for cell in row])
    result[sheet_name] = rows

with open("db/excel_full.json", "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

print(f"시트 수: {len(result)}")
for name, rows in result.items():
    print(f"  {name}: {len(rows)}행")
